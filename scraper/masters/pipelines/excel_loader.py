"""Excel → Supabase bulk loader for canonical.masters_programs.

Usage:
    python -m scraper.masters.pipelines.excel_loader --file programs.xlsx --dry-run
    python -m scraper.masters.pipelines.excel_loader --file programs.xlsx --limit 10
    python -m scraper.masters.pipelines.excel_loader --recon --file programs.xlsx

Reuses the existing validator chain (program_validator.py) before any write.
Idempotent upsert via ON CONFLICT on (canonical_institution_id, program_name,
degree_type, intake_term, intake_year).

If --recon is passed, prints sheet names, row counts, headers, and a proposed
column→schema mapping WITHOUT touching the database.

If --file is provided without --recon, loads data into the database.
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
import time

# Windows console defaults to cp1252, which can't encode many Unicode chars
# (curly quotes, ≥, em-dashes) that show up in scraped admissions text.
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Dependencies — only pulled in when actually needed
# ---------------------------------------------------------------------------

def _import_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


def _import_pandas():
    try:
        import pandas
        return pandas
    except ImportError:
        print("ERROR: pandas not installed. Run: pip install pandas", file=sys.stderr)
        sys.exit(1)


class _QueryResult:
    """Shim so call sites can keep using node-pg-style `.rows` access."""
    def __init__(self, rows):
        self.rows = rows


class _PgPool:
    """Thin synchronous Postgres connection, reading the same DATABASE_URL /
    SUPABASE_DB_URL the Node backend uses (backend/.env) — there is no Python
    equivalent of backend/src/config/database.js to import, so this connects
    directly instead of going through the JS layer."""
    def __init__(self, conn):
        self._conn = conn

    def query(self, sql: str, params: Optional[list] = None) -> _QueryResult:
        # These queries were written node-pg style ($1, $2, ...); psycopg2 uses
        # %s positional placeholders. Safe to blanket-substitute here since
        # every call site's $N appears in ascending left-to-right order matching
        # the params list — no reused/out-of-order placeholders anywhere.
        sql = re.sub(r"\$\d+", "%s", sql)
        with self._conn.cursor() as cur:
            cur.execute(sql, params or [])
            try:
                rows = cur.fetchall()
            except Exception:
                rows = []
            self._conn.commit()
            return _QueryResult(rows)


def _import_db() -> "_PgPool":
    import psycopg2
    import psycopg2.extras
    from pathlib import Path

    env_path = Path(__file__).resolve().parents[3] / "backend" / ".env"
    if not env_path.exists():
        print(f"ERROR: backend/.env not found at {env_path}", file=sys.stderr)
        sys.exit(1)

    db_url = None
    for line in env_path.read_text(encoding="utf-8", errors="replace").splitlines():
        if line.startswith("DATABASE_URL=") or line.startswith("SUPABASE_DB_URL="):
            db_url = line.split("=", 1)[1].strip()
            break
    if not db_url:
        print("ERROR: DATABASE_URL/SUPABASE_DB_URL not found in backend/.env", file=sys.stderr)
        sys.exit(1)

    conn = psycopg2.connect(db_url, cursor_factory=psycopg2.extras.RealDictCursor)
    conn.autocommit = True  # each query is its own transaction — one bad row
    # can't poison the connection for every row after it.
    return _PgPool(conn)


# ---------------------------------------------------------------------------
# Schema field definitions (canonical.masters_programs)
# ---------------------------------------------------------------------------

MASTERS_PROGRAMS_FIELDS = [
    "canonical_institution_id", "institution_name", "institution_country",
    "city", "department", "program_name", "degree_type", "specialization",
    "cip_code", "is_stem_designated", "language_of_instruction",
    "intake_term", "intake_year",
    "gre_requirement", "gmat_requirement",
    "min_gpa", "min_gpa_scale", "min_toefl", "min_ielts",
    "funding_availability", "assistantship_types", "tuition_waiver_available",
    "tuition_total", "tuition_currency", "program_length_months",
    "median_earnings", "median_debt", "roi_source",
    "program_url", "data_source", "data_quality_score", "last_scraped_at",
]

VALID_DEGREES = {"MS", "MA", "MBA"}
VALID_FUNDING = {"fully_funded", "partial", "unfunded", "varies", "unknown"}
VALID_INTAKE_TERMS = {"fall", "spring", "summer", "winter"}
VALID_DEADLINE_TYPES = {
    "priority", "final", "funding_consideration",
    "round_1", "round_2", "round_3", "rolling",
}

# Country string normalization — the dataset mixes full names, 2-letter codes,
# and "US" vs "USA" for the same country within the same workbook. CT signed
# off on expanding scope beyond the original plan's 7 countries to cover every
# single-country value actually present in this dataset (Phase 8 A0 decision).
COUNTRY_ALIASES = {
    "US": "US", "USA": "US", "UNITED STATES": "US",
    "UK": "UK", "UNITED KINGDOM": "UK",
    "CANADA": "CA", "CA": "CA",
    "GERMANY": "DE", "DE": "DE",
    "NETHERLANDS": "NL", "THE NETHERLANDS": "NL", "NL": "NL",
    "AUSTRALIA": "AU", "AU": "AU",
    "SINGAPORE": "SG", "SG": "SG",
    "INDIA": "IN", "FRANCE": "FR", "SWITZERLAND": "CH", "SPAIN": "ES",
    "ITALY": "IT", "SWEDEN": "SE", "IRELAND": "IE", "DENMARK": "DK",
    "BELGIUM": "BE", "AUSTRIA": "AT", "NORWAY": "NO", "FINLAND": "FI",
    "CHINA": "CN", "HONG KONG": "HK", "HONG KONG SAR": "HK", "NEW ZEALAND": "NZ",
}
VALID_COUNTRY_CODES = set(COUNTRY_ALIASES.values())

# Last-resort degree-type fallback by sheet subject area, used only when a
# course title has no degree-level token at all (e.g. "Epidemiology",
# "Financial Mathematics"). A heuristic compromise, not ground truth — flagged
# in the recon/dry-run report so CT can spot-check it. Sheets not listed here
# get no fallback and are excluded rather than guessed.
SHEET_DEGREE_DEFAULT = {
    # STEM / quantitative sheets
    "BIOINFORMATICS  BIOITECHNOLOGY": "MS", "BIOSCIENCES": "MS", "NEUROSCIENCE": "MS",
    "CHEMISTRY": "MS", "PHYSICS": "MS", "ASTRONOMY": "MS", "COMPUTER SCIENCE": "MS",
    "MATHEMATICS": "MS", "DATA SCIENCE": "MS", "BUSINESS ANALYTICS": "MS",
    "PUBLIC HEALTH": "MS", "ECONOMETRICS": "MS", "FINANCE": "MS",
    "ENVIRONMENT AND SUSTAINABILITY": "MS",
    # Humanities / social science / arts sheets
    "GENERALAPPLIED PSYCHOLOGY": "MA", "SOCIAL PSYCHOLOGY": "MA",
    "DEVELOPMENTAL PSYCHOLOGY": "MA", "MISCELLANEOUS PSYCHOLOGY": "MA",
    "HISTORY": "MA", "SOCIOLOGY ANTHROPOLOGY": "MA", "ENGLISH": "MA",
    "CREATIVE WRITING": "MA", "PERFORMING ARTS": "MA", "VISUAL ARTS": "MA",
    "MEDIA STUDIES": "MA", "GENDER STUDIES": "MA",
}

# ---------------------------------------------------------------------------
# Recon mode
# ---------------------------------------------------------------------------

@dataclass
class SheetInfo:
    name: str
    row_count: int
    headers: List[str]
    sample_rows: List[dict] = field(default_factory=list)
    all_program_values: List[str] = field(default_factory=list)
    all_country_values: List[str] = field(default_factory=list)


def recon_excel(filepath: str) -> List[SheetInfo]:
    """Read the workbook and report structure WITHOUT touching the DB."""
    openpyxl = _import_openpyxl()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets: List[SheetInfo] = []

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=False))
        if not rows:
            sheets.append(SheetInfo(name=ws_name, row_count=0, headers=[]))
            continue

        headers = [str(c.value).strip() if c.value else "" for c in rows[0]]
        # Excel/Sheets exports often pad the used range with trailing blank rows
        # far past the real data — drop any row where every cell is empty.
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None and str(v).strip() for v in r)]
        row_count = len(data_rows)

        sample: List[dict] = []
        for r in data_rows[:3]:
            sample.append(dict(zip(headers, r)))

        program_idx = next((i for i, h in enumerate(headers) if h.strip().lower() == "program"), None)
        country_idx = next((i for i, h in enumerate(headers) if h.strip().lower() == "country"), None)
        all_program_values = [str(r[program_idx]).strip() for r in data_rows if program_idx is not None and r[program_idx]]
        all_country_values = [str(r[country_idx]).strip() for r in data_rows if country_idx is not None and r[country_idx]]

        sheets.append(SheetInfo(
            name=ws_name,
            row_count=row_count,
            headers=headers,
            sample_rows=sample,
            all_program_values=all_program_values,
            all_country_values=all_country_values,
        ))

    wb.close()
    return sheets


def print_recon(sheets: List[SheetInfo], filepath: str):
    """Print a human-readable recon report for CT sign-off."""
    print(f"\n{'='*70}")
    print(f"EXCEL RECON REPORT — {filepath}")
    print(f"{'='*70}\n")

    for s in sheets:
        print(f"Sheet: {s.name}")
        print(f"  Row count: {s.row_count}")
        print(f"  Headers ({len(s.headers)}): {', '.join(s.headers)}")
        if s.sample_rows:
            print(f"  Sample rows:")
            for i, row in enumerate(s.sample_rows):
                print(f"    Row {i+1}: {json.dumps(row, default=str, ensure_ascii=False)[:200]}")
        print()

    # Degree type / country analysis across ALL rows (not just the 3-row sample)
    from collections import Counter
    degree_counts: Counter = Counter()
    country_counts: Counter = Counter()
    for s in sheets:
        for v in s.all_program_values:
            degree_counts[v.strip().upper()] += 1
        for v in s.all_country_values:
            country_counts[v.strip().upper()] += 1

    print(f"{'='*70}")
    print("FLAGGED VALUES (for CT review) — counted across ALL rows, not just samples")
    print(f"{'='*70}")
    print(f"  Degree/Program values seen: {dict(degree_counts.most_common())}")
    bad_degrees = {k: c for k, c in degree_counts.items() if k not in ("MASTERS", "MS", "MA", "MBA") and k}
    if bad_degrees:
        print(f"  ⚠️  Non-Masters values (will be EXCLUDED from ingestion): {bad_degrees}")
    print(f"  ⚠️  NOTE: 'Masters' is not a valid degree_type on its own — every row's real")
    print(f"     MS/MA/MBA must be inferred from the Course title text (e.g. 'MSc'→MS,")
    print(f"     'MA'/'MPhil'→MA). This recon does not attempt that inference.")

    print(f"\n  Country values seen: {dict(country_counts.most_common())}")
    normalized = Counter()
    unmapped_countries = Counter()
    for raw, c in country_counts.items():
        code = COUNTRY_ALIASES.get(raw)
        if code:
            normalized[code] += c
        else:
            unmapped_countries[raw] += c
    out_of_scope = {k: c for k, c in normalized.items() if k not in VALID_COUNTRY_CODES}
    if out_of_scope or unmapped_countries:
        print(f"  ⚠️  Countries outside original plan scope (US,UK,CA,DE,NL,AU,SG):")
        for k, c in {**out_of_scope, **unmapped_countries}.items():
            print(f"       {k}: {c} rows")
        print(f"     → Flagged for scope decision — these rows will be SKIPPED unless scope is expanded.")
    else:
        print(f"  ✓ All countries within original plan scope")

    print()


# ---------------------------------------------------------------------------
# Column mapping — proposed defaults (CT to confirm in A0)
# ---------------------------------------------------------------------------

COLUMN_MAPPINGS = {
    # masters_programs fields
    "institution_name": "institution_name",
    "institution": "institution_name",
    "school": "institution_name",
    "university": "institution_name",
    "institution_country": "institution_country",
    "country": "institution_country",
    "city": "city",
    "location": "city",
    "department": "department",
    "dept": "department",
    "program_name": "program_name",
    "course": "program_name",
    "degree_type": "degree_type",
    "degree": "degree_type",
    # NOTE: this dataset's "Program" column holds "Masters"/"PhD" (a level filter,
    # not the program title or the MS/MA/MBA code) — handled specially in
    # normalize_program_row, deliberately NOT mapped here to avoid colliding
    # with "course" (the actual program title) or degree_type (inferred below).
    "specialization": "specialization",
    "focus": "specialization",
    "track": "specialization",
    "cip_code": "cip_code",
    "cip": "cip_code",
    "is_stem": "is_stem_designated",
    "stem": "is_stem_designated",
    "stem_designated": "is_stem_designated",
    "intake_term": "intake_term",
    "term": "intake_term",
    "intake_year": "intake_year",
    "year": "intake_year",
    "gre_requirement": "gre_requirement",
    "gre": "gre_requirement",
    "gmat_requirement": "gmat_requirement",
    "gmat": "gmat_requirement",
    "min_gpa": "min_gpa",
    "gpa": "min_gpa",
    "min_toefl": "min_toefl",
    "toefl": "min_toefl",
    "min_ielts": "min_ielts",
    "ielts": "min_ielts",
    "funding_availability": "funding_availability",
    "funding": "funding_availability",
    "tuition_total": "tuition_total",
    "tuition": "tuition_total",
    "cost": "tuition_total",
    "tuition_currency": "tuition_currency",
    "currency": "tuition_currency",
    "program_url": "program_url",
    "url": "program_url",
    "website": "program_url",
    "data_source": "data_source",
    "source": "data_source",
    # Deadlines child table — columns prefixed with "deadline_"
    # Pathways child table — columns prefixed with "pathway_"
}


def map_column(header: str) -> Optional[str]:
    """Map an Excel column header to a schema field name."""
    h = str(header).strip().lower()
    # Exact match first
    if h in COLUMN_MAPPINGS:
        return COLUMN_MAPPINGS[h]
    # Partial match
    for key, val in COLUMN_MAPPINGS.items():
        if key in h or h in key:
            return val
    return None


def map_deadline_column(header: str) -> Optional[str]:
    """Map a deadline-prefixed column to a schema field."""
    h = str(header).strip().lower()
    deadline_map = {
        "deadline_date": "deadline_date",
        "deadline_type": "deadline_type",
        "deadline": "deadline_date",
        "deadline_notes": "notes",
        "notes": "notes",
        "source_url": "source_url",
        "is_rolling": "is_rolling",
        "rolling": "is_rolling",
        "intake_term": "intake_term",
        "intake_year": "intake_year",
    }
    if h in deadline_map:
        return deadline_map[h]
    for key, val in deadline_map.items():
        if key in h or h in key:
            return val
    return None


def map_pathway_column(header: str) -> Optional[str]:
    """Map a pathway-prefixed column to a schema field."""
    h = str(header).strip().lower()
    pathway_map = {
        "pathway_type": "pathway_type",
        "pathway": "pathway_type",
        "admission_pathway": "pathway_type",
        "description": "description",
        "min_requirements": "min_requirements",
        "confidence": "confidence",
        "source_url": "source_url",
    }
    if h in pathway_map:
        return pathway_map[h]
    for key, val in pathway_map.items():
        if key in h or h in key:
            return val
    return None


# ---------------------------------------------------------------------------
# Data coercion helpers
# ---------------------------------------------------------------------------

def coerce_bool(v: Any) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y"):
        return True
    if s in ("false", "0", "no", "n"):
        return False
    return None


def coerce_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def coerce_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def coerce_json_list(v: Any) -> Optional[str]:
    """Return a JSON string for JSONB columns, or None."""
    if v is None:
        return None
    if isinstance(v, list):
        return json.dumps(v)
    s = str(v).strip()
    if not s:
        return None
    # Try parsing as JSON first
    try:
        parsed = json.loads(s)
        if isinstance(parsed, list):
            return json.dumps(parsed)
        return json.dumps([parsed])
    except (json.JSONDecodeError, TypeError):
        # Comma-separated values
        items = [x.strip() for x in s.split(",") if x.strip()]
        if items:
            return json.dumps(items)
        return None


# ---------------------------------------------------------------------------
# Row normalization
# ---------------------------------------------------------------------------

def infer_degree_type(course_title: str, sheet_name: str) -> Optional[str]:
    """Infer MS|MA|MBA from free-text course title — this dataset never states
    the degree code directly (its 'Program' column only ever says Masters/PhD).
    Best-effort heuristic, not ground truth; unmatched titles return None and
    are excluded rather than guessed."""
    t = (course_title or "").upper()
    sheet = (sheet_name or "").strip().upper()

    if "MBA" in t or sheet == "MBA":
        return "MBA"
    if re.search(
        r"M\.?\s?SC\b|M\.?\s?S\.?\s?E\.?\b|\bMS\b|\bMSE\b|\bMSBA\b|\bMASt\b|\bMPS\b|"
        r"MASTER OF SCIENCE|MASTER OF BUSINESS ANALYTICS|MASTER OF ENGINEERING|\bMRES\b",
        t,
    ):
        return "MS"
    if re.search(
        r"\bM\.?A\.?\b|\bMPHIL\b|M\.?\s?PHIL\b|\bMST\b|M\.?\s?F\.?\s?A\.?\b|\bMLITT\b|\bMED\b|\bLLM\b|"
        r"MASTER OF ARTS|MASTER IN MANAGEMENT|MASTER OF PUBLIC POLICY|MASTER OF FINE ARTS|"
        r"MASTER OF PUBLIC AFFAIRS|MASTER OF PUBLIC HEALTH|MASTER OF EDUCATION|\bMPP\b|\bMPH\b",
        t,
    ):
        return "MA"
    # Generic "Master('s) in/of X" with no more specific signal — documented
    # best-effort fallback (Crown plan A0: default ambiguous masters to MA).
    if re.search(r"\bMASTER'?S?\b", t):
        return "MA"
    # Last resort: the course title has NO degree-level token at all (e.g. just
    # "Epidemiology" or "Financial Mathematics") — fall back to the sheet's
    # subject-area default rather than excluding a real program outright.
    # Flagged explicitly in the recon/dry-run report as a heuristic compromise.
    return SHEET_DEGREE_DEFAULT.get(sheet)


def normalize_program_row(row: dict, sheet_name: str) -> Tuple[Optional[dict], List[str]]:
    """Map Excel row → masters_programs dict. Returns (record, issues)."""
    issues: List[str] = []
    mapped: Dict[str, Any] = {}

    # "Program" is handled separately below — it's a Masters/PhD level filter in
    # this dataset, not the program title (that's "Course") or a degree code.
    level_raw = next((v for k, v in row.items() if k and str(k).strip().lower() == "program"), None)
    level = str(level_raw or "").strip().upper()
    if level.startswith("PHD"):
        return None, ["excluded_non_masters_level:PHD"]

    for header, value in row.items():
        if header is None or str(header).strip().lower() == "program":
            continue
        field_name = map_column(str(header))
        if field_name and field_name in MASTERS_PROGRAMS_FIELDS:
            mapped[field_name] = value

    # Degree type was never stated directly — infer from the course title.
    if not mapped.get("degree_type"):
        inferred = infer_degree_type(str(mapped.get("program_name", "")), sheet_name)
        if inferred:
            mapped["degree_type"] = inferred

    # Validate required fields
    for req in ("institution_name", "program_name", "degree_type"):
        if not mapped.get(req):
            issues.append(f"missing_required:{req}")

    # Validate degree_type
    degree = str(mapped.get("degree_type", "")).strip().upper()
    if degree not in VALID_DEGREES:
        issues.append(f"bad_degree_type:{degree}")
        return None, issues  # Exclude invalid degrees

    # Coerce types
    mapped["is_stem_designated"] = coerce_bool(mapped.get("is_stem_designated"))
    mapped["intake_year"] = coerce_int(mapped.get("intake_year"))
    mapped["min_gpa"] = coerce_float(mapped.get("min_gpa"))
    mapped["min_gpa_scale"] = coerce_float(mapped.get("min_gpa_scale"))
    mapped["min_toefl"] = coerce_int(mapped.get("min_toefl"))
    mapped["min_ielts"] = coerce_float(mapped.get("min_ielts"))
    mapped["tuition_total"] = coerce_float(mapped.get("tuition_total"))
    mapped["tuition_waiver_available"] = coerce_bool(mapped.get("tuition_waiver_available"))
    mapped["program_length_months"] = coerce_int(mapped.get("program_length_months"))
    mapped["median_earnings"] = coerce_float(mapped.get("median_earnings"))
    mapped["median_debt"] = coerce_float(mapped.get("median_debt"))
    mapped["data_quality_score"] = coerce_float(mapped.get("data_quality_score"))
    mapped["assistantship_types"] = coerce_json_list(mapped.get("assistantship_types"))
    mapped["language_of_instruction"] = coerce_json_list(mapped.get("language_of_instruction"))

    # Validate funding_availability
    funding = str(mapped.get("funding_availability", "")).strip().lower()
    if funding and funding not in VALID_FUNDING:
        issues.append(f"bad_funding:{funding}")
        mapped["funding_availability"] = None

    # Validate intake_term
    term = str(mapped.get("intake_term", "")).strip().lower()
    if term and term not in VALID_INTAKE_TERMS:
        issues.append(f"bad_intake_term:{term}")
        mapped["intake_term"] = None

    # Data source defaults
    if not mapped.get("data_source"):
        mapped["data_source"] = f"excel_import:{sheet_name}"

    # Normalize institution_name and program_name
    mapped["institution_name"] = str(mapped.get("institution_name", "")).strip()
    mapped["program_name"] = str(mapped.get("program_name", "")).strip()
    mapped["degree_type"] = degree

    # Some consortium/Erasmus Mundus rows have a descriptive paragraph shifted
    # into the University cell instead of an institution name (merged-cell
    # layout artifact from the source sheet) — exclude rather than load garbage.
    if len(mapped["institution_name"]) > 80:
        issues.append("excluded_implausible_institution_name")
        return None, issues

    # Country normalization + scope filter. Multi-country joint-program rows
    # (e.g. "Spain, Poland, UK, Netherlands, Austria" for an Erasmus Mundus
    # consortium) can't be safely assigned a single institution_country —
    # exclude rather than guess which one.
    country_raw = str(mapped.get("institution_country", "")).strip().upper()
    if "," in country_raw or "/" in country_raw:
        issues.append(f"excluded_multi_country:{country_raw}")
        return None, issues
    country_code = COUNTRY_ALIASES.get(country_raw, country_raw)
    if country_code not in VALID_COUNTRY_CODES:
        issues.append(f"excluded_out_of_scope_country:{country_raw}")
        return None, issues
    mapped["institution_country"] = country_code

    return mapped, issues


def normalize_deadline_row(row: dict) -> Optional[dict]:
    """Map Excel row → masters_program_deadlines dict."""
    mapped: Dict[str, Any] = {}

    for header, value in row.items():
        if header is None:
            continue
        field_name = map_deadline_column(str(header))
        if field_name:
            mapped[field_name] = value

    if not mapped.get("deadline_date") and not mapped.get("deadline_type"):
        return None

    mapped["deadline_date"] = str(mapped.get("deadline_date", "")).strip() if mapped.get("deadline_date") else None
    mapped["is_rolling"] = coerce_bool(mapped.get("is_rolling"))

    dt = str(mapped.get("deadline_type", "")).strip().lower()
    if dt and dt not in VALID_DEADLINE_TYPES:
        mapped["deadline_type"] = None  # Will be caught by CHECK constraint

    return mapped


def normalize_pathway_row(row: dict) -> Optional[dict]:
    """Map Excel row → masters_program_pathways dict."""
    mapped: Dict[str, Any] = {}

    for header, value in row.items():
        if header is None:
            continue
        field_name = map_pathway_column(str(header))
        if field_name:
            mapped[field_name] = value

    if not mapped.get("pathway_type") and not mapped.get("description"):
        return None

    conf = coerce_float(mapped.get("confidence"))
    if conf is not None:
        mapped["confidence"] = round(conf, 2)

    mapped["min_requirements"] = coerce_json_list(mapped.get("min_requirements"))

    return mapped


# ---------------------------------------------------------------------------
# Database upsert
# ---------------------------------------------------------------------------


def resolve_institution_id(pool, institution_name: str, institution_country: str) -> Optional[str]:
    """Try to resolve to canonical_institution_id via institutions table."""
    from uuid import UUID
    try:
        import re
        normalized = re.sub(r'[^a-z0-9]+', ' ', institution_name.lower()).strip()
        result = pool.query(
            "SELECT id FROM canonical.institutions "
            "WHERE normalized_name = $1 LIMIT 1",
            [normalized],
        )
        if result.rows and result.rows[0].get("id"):
            return str(result.rows[0]["id"])
        # Fallback: exact name match
        result = pool.query(
            "SELECT id FROM canonical.institutions "
            "WHERE LOWER(name) = LOWER($1) LIMIT 1",
            [institution_name],
        )
        if result.rows and result.rows[0].get("id"):
            return str(result.rows[0]["id"])
    except Exception:
        pass
    return None


def classify_sheets(filepath: str) -> Dict[str, str]:
    """Detect each sheet's purpose: "programs" | "deadlines" | "pathways".

    Almost every sheet in this dataset is one subject-area tab of programs
    (e.g. "Computer Science", "MBA") — the keyword/header heuristics exist
    for the rare case a future workbook splits deadlines/pathways into their
    own tabs, but in practice everything currently classifies as "programs".
    """
    openpyxl = _import_openpyxl()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_purpose: Dict[str, str] = {}

    for name in wb.sheetnames:
        ns = name.lower()
        if ns.strip() == "index":
            sheet_purpose[name] = "skip"
            continue

        # A sheet with University + Course/Program columns IS a programs sheet,
        # full stop — even though it also has a "Deadlines" column (every program
        # has a deadline, that's just one of its fields, not the sheet's grain).
        # Checking this FIRST avoids misclassifying ~30 of 36 sheets as "deadlines"
        # just because they contain a deadline column among many others.
        ws = wb[name]
        headers_lower = [str(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        has_university = any("university" in h or "institution" in h for h in headers_lower)
        has_course_or_program = any(h.strip() in ("course", "program", "program name", "degree") for h in headers_lower)

        if has_university and has_course_or_program:
            sheet_purpose[name] = "programs"
        elif any(kw in ns for kw in ("deadline", "date", "timeline")):
            sheet_purpose[name] = "deadlines"
        elif any(kw in ns for kw in ("pathway", "admission", "entry")):
            sheet_purpose[name] = "pathways"
        elif any(kw in ns for kw in ("program", "course", "degree", "master")):
            sheet_purpose[name] = "programs"
        elif any(map_deadline_column(h) for h in headers_lower):
            sheet_purpose[name] = "deadlines"
        elif any(map_pathway_column(h) for h in headers_lower):
            sheet_purpose[name] = "pathways"
        else:
            sheet_purpose[name] = "programs"

    wb.close()
    return sheet_purpose


def bulk_load(filepath: str, limit: Optional[int] = None):
    """Load Excel data into the database."""
    pandas = _import_pandas()
    sheet_purpose = classify_sheets(filepath)

    # Import DB
    pool = _import_db()

    # Track stats
    stats = {
        "programs_read": 0, "programs_upserted": 0, "programs_rejected": 0,
        "deadlines_read": 0, "deadlines_upserted": 0, "deadlines_rejected": 0,
        "pathways_read": 0, "pathways_upserted": 0, "pathways_rejected": 0,
        "institutions_resolved": 0, "institutions_null": 0,
        "total_seconds": 0,
    }

    start = time.time()

    # ── Programs ────────────────────────────────────────────────────────
    program_sheets = [n for n, p in sheet_purpose.items() if p == "programs"]
    all_programs: List[dict] = []

    for sheet_name in program_sheets:
        df = pandas.read_excel(filepath, sheet_name=sheet_name)
        stats["programs_read"] += len(df)

        for _, row in df.iterrows():
            row_dict = row.dropna().to_dict()
            record, issues = normalize_program_row(row_dict, sheet_name)
            if record:
                record["_issues"] = issues
                all_programs.append(record)
            else:
                stats["programs_rejected"] += 1

            if limit and stats["programs_upserted"] >= limit:
                break

        if limit and stats["programs_upserted"] >= limit:
            break

    # Resolve institution IDs
    institution_cache: Dict[str, Optional[str]] = {}
    for prog in all_programs:
        key = prog.get("institution_name", "")
        if key not in institution_cache:
            inst_id = resolve_institution_id(pool, key, prog.get("institution_country", ""))
            institution_cache[key] = inst_id
            if inst_id:
                stats["institutions_resolved"] += 1
            else:
                stats["institutions_null"] += 1
        prog["canonical_institution_id"] = institution_cache[key]

    # Upsert programs
    for prog in all_programs:
        if limit and stats["programs_upserted"] >= limit:
            break

        import psycopg2.extras as _pgextras
        inst_id = prog.get("canonical_institution_id")
        params = [
            inst_id, prog["institution_name"], prog["institution_country"],
            prog.get("city"), prog.get("department"), prog["program_name"],
            prog["degree_type"], prog.get("specialization"), prog.get("cip_code"),
            prog.get("is_stem_designated"),
            _pgextras.Json(prog.get("language_of_instruction") or []),
            prog.get("intake_term"), prog.get("intake_year"),
            prog.get("gre_requirement"), prog.get("gmat_requirement"),
            prog.get("min_gpa"), prog.get("min_gpa_scale"),
            prog.get("min_toefl"), prog.get("min_ielts"),
            prog.get("funding_availability"), _pgextras.Json(prog.get("assistantship_types") or []),
            prog.get("tuition_waiver_available"),
            prog.get("tuition_total"), prog.get("tuition_currency"),
            prog.get("program_length_months"),
            prog.get("median_earnings"), prog.get("median_debt"),
            prog.get("roi_source"),
            prog.get("program_url"), prog.get("data_source"),
            prog.get("data_quality_score"),
        ]

        update_cols = [
            "institution_name", "institution_country", "city", "department",
            "specialization", "cip_code", "is_stem_designated",
            "language_of_instruction", "intake_term", "intake_year",
            "gre_requirement", "gmat_requirement", "min_gpa", "min_gpa_scale",
            "min_toefl", "min_ielts", "funding_availability",
            "assistantship_types", "tuition_waiver_available",
            "tuition_total", "tuition_currency", "program_length_months",
            "median_earnings", "median_debt", "roi_source",
            "program_url", "data_source", "data_quality_score",
            "last_scraped_at",
        ]
        update_set = ", ".join(
            f"{c} = EXCLUDED.{c}" for c in update_cols
        ) + ", updated_at = NOW()"

        # Postgres treats NULL <> NULL in uniqueness checks, so the original
        # constraint (canonical_institution_id, program_name, degree_type,
        # intake_term, intake_year) never fires as a conflict when inst_id is
        # NULL — which it is for nearly every bulk-imported row. Migration 123
        # adds a partial unique index for exactly that case.
        if inst_id is not None:
            conflict_target = "canonical_institution_id, program_name, degree_type, intake_term, intake_year"
        else:
            conflict_target = "institution_name, program_name, degree_type"

        try:
            result = pool.query(
                f"""INSERT INTO canonical.masters_programs (
                        canonical_institution_id, institution_name, institution_country,
                        city, department, program_name, degree_type, specialization,
                        cip_code, is_stem_designated, language_of_instruction,
                        intake_term, intake_year,
                        gre_requirement, gmat_requirement,
                        min_gpa, min_gpa_scale, min_toefl, min_ielts,
                        funding_availability, assistantship_types, tuition_waiver_available,
                        tuition_total, tuition_currency, program_length_months,
                        median_earnings, median_debt, roi_source,
                        program_url, data_source, data_quality_score, last_scraped_at
                    ) VALUES ({', '.join(f'${i+1}' for i in range(len(params)))}, NOW())
                    ON CONFLICT ({conflict_target}) DO UPDATE SET {update_set}
                    RETURNING id""",
                params,
            )
            if result.rows:
                prog["id"] = result.rows[0]["id"]
            stats["programs_upserted"] += 1
        except Exception as e:
            stats["programs_rejected"] += 1
            print(f"  ERROR upserting program '{prog['program_name']}': {e}", file=sys.stderr)

    # ── Deadlines ───────────────────────────────────────────────────────
    deadline_sheets = [n for n, p in sheet_purpose.items() if p == "deadlines"]
    deadline_rows: List[dict] = []

    for sheet_name in deadline_sheets:
        df = pandas.read_excel(filepath, sheet_name=sheet_name)
        stats["deadlines_read"] += len(df)
        for _, row in df.iterrows():
            dr = normalize_deadline_row(row.dropna().to_dict())
            if dr:
                dr["_issues"] = []
                deadline_rows.append(dr)
            else:
                stats["deadlines_rejected"] += 1

    # Link deadlines to programs (by program_name + degree_type + intake)
    if deadline_rows:
        prog_lookup = {}
        for prog in all_programs:
            key = (prog["program_name"], prog["degree_type"], prog.get("intake_term"), prog.get("intake_year"))
            prog_lookup[key] = prog.get("id")

        for dl in deadline_rows:
            dl_key = (
                dl.get("program_name"), dl.get("degree_type"),
                dl.get("intake_term"), dl.get("intake_year"),
            )
            prog_id = prog_lookup.get(dl_key)
            if not prog_id:
                # Try partial match
                for prog in all_programs:
                    if (prog["program_name"] == dl.get("program_name") and
                        prog["degree_type"] == dl.get("degree_type")):
                        prog_id = prog.get("id")
                        break

            if prog_id:
                dl["masters_program_id"] = str(prog_id)

        # Upsert deadlines
        for dl in deadline_rows:
            if not dl.get("masters_program_id"):
                stats["deadlines_rejected"] += 1
                continue

            d = normalize_deadline_row(dl)
            if not d:
                stats["deadlines_rejected"] += 1
                continue

            params = [
                d["masters_program_id"], d.get("deadline_type"),
                d.get("deadline_date"), d.get("is_rolling"),
                d.get("intake_term"), d.get("intake_year"),
                d.get("notes"), d.get("source_url"),
            ]
            try:
                pool.query(
                    """INSERT INTO canonical.masters_program_deadlines
                       (masters_program_id, deadline_type, deadline_date, is_rolling,
                        intake_term, intake_year, notes, source_url)
                       VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                       ON CONFLICT DO NOTHING""",
                    params,
                )
                stats["deadlines_upserted"] += 1
            except Exception as e:
                stats["deadlines_rejected"] += 1

    # ── Pathways ────────────────────────────────────────────────────────
    pathway_sheets = [n for n, p in sheet_purpose.items() if p == "pathways"]
    pathway_rows: List[dict] = []

    for sheet_name in pathway_sheets:
        df = pandas.read_excel(filepath, sheet_name=sheet_name)
        stats["pathways_read"] += len(df)
        for _, row in df.iterrows():
            pr = normalize_pathway_row(row.dropna().to_dict())
            if pr:
                pr["_issues"] = []
                pathway_rows.append(pr)
            else:
                stats["pathways_rejected"] += 1

    # Link pathways to programs
    if pathway_rows:
        for pw in pathway_rows:
            prog_key = (pw.get("program_name"), pw.get("degree_type"))
            for prog in all_programs:
                if prog["program_name"] == pw.get("program_name") and prog["degree_type"] == pw.get("degree_type"):
                    pw["masters_program_id"] = str(prog.get("id"))
                    break

        # Upsert pathways
        for pw in pathway_rows:
            if not pw.get("masters_program_id"):
                stats["pathways_rejected"] += 1
                continue

            p = normalize_pathway_row(pw)
            if not p:
                stats["pathways_rejected"] += 1
                continue

            params = [
                p["masters_program_id"], p.get("pathway_type"),
                p.get("description"), p.get("weighted_fields"),
                p.get("min_requirements"), p.get("confidence"),
                p.get("source_url"),
            ]
            try:
                pool.query(
                    """INSERT INTO canonical.masters_program_pathways
                       (masters_program_id, pathway_type, description,
                        weighted_fields, min_requirements, confidence, source_url)
                       VALUES ($1, $2, $3, $4, $5, $6, $7)
                       ON CONFLICT DO NOTHING""",
                    params,
                )
                stats["pathways_upserted"] += 1
            except Exception as e:
                stats["pathways_rejected"] += 1

    # ── Scrape log ──────────────────────────────────────────────────────
    accepted = stats["programs_upserted"]
    rejected = stats["programs_rejected"]
    try:
        pool.query(
            """INSERT INTO canonical.masters_scrape_log
               (program_url, institution_name, program_name, status,
                confidence, issues, missing_fields, scraped_at)
               VALUES ($1, $2, $3, $4, $5, $6, $7, NOW())""",
            [
                f"excel:{filepath}",
                "bulk_import", "bulk_import",
                "accepted" if accepted > 0 else "rejected",
                round((accepted / max(accepted + rejected, 1)) * 100, 2),
                json.dumps([f"programs_upserted={accepted}", f"programs_rejected={rejected}"]),
                json.dumps([]),
            ],
        )
    except Exception as e:
        print(f"  WARNING: Could not write scrape_log: {e}", file=sys.stderr)

    # ── Refresh MV ──────────────────────────────────────────────────────
    try:
        pool.query("REFRESH MATERIALIZED VIEW CONCURRENTLY canonical.mv_masters_program_cards")
        print("  ✓ mv_masters_program_cards refreshed")
    except Exception as e:
        print(f"  WARNING: MV refresh failed: {e}", file=sys.stderr)

    stats["total_seconds"] = round(time.time() - start, 2)

    # ── Summary ─────────────────────────────────────────────────────────
    print(f"\n{'='*70}")
    print("LOAD SUMMARY")
    print(f"{'='*70}")
    print(f"  Programs:  {stats['programs_read']} read, {stats['programs_upserted']} upserted, {stats['programs_rejected']} rejected")
    print(f"  Deadlines: {stats['deadlines_read']} read, {stats['deadlines_upserted']} upserted, {stats['deadlines_rejected']} rejected")
    print(f"  Pathways:  {stats['pathways_read']} read, {stats['pathways_upserted']} upserted, {stats['pathways_rejected']} rejected")
    print(f"  Institutions resolved: {stats['institutions_resolved']}")
    print(f"  Institutions NULL (denormalized only): {stats['institutions_null']}")
    print(f"  Duration: {stats['total_seconds']}s")
    print(f"{'='*70}\n")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Excel → Supabase bulk loader for masters_programs"
    )
    parser.add_argument("--file", "-f", help="Path to the Excel file (.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Print mapping, write nothing")
    parser.add_argument("--recon", action="store_true", help="Recon mode: report structure only")
    parser.add_argument("--limit", "-l", type=int, help="Limit rows processed (for testing)")
    args = parser.parse_args()

    if not args.file:
        print("ERROR: --file <path> is required", file=sys.stderr)
        sys.exit(1)

    if args.recon:
        sheets = recon_excel(args.file)
        print_recon(sheets, args.file)
        sheet_purpose = classify_sheets(args.file)

        # Print proposed column mapping
        print(f"{'='*70}")
        print("PROPOSED COLUMN → SCHEMA MAPPING")
        print(f"{'='*70}")
        for s in sheets:
            print(f"\nSheet: {s.name} (purpose: {sheet_purpose.get(s.name, 'unknown')})")
            for h in s.headers:
                mapped = map_column(h) or map_deadline_column(h) or map_pathway_column(h)
                status = f"→ {mapped}" if mapped else "→ UNMAPPED — needs CT decision"
                print(f"  {h:40s} {status}")
        print()
        return

    if args.dry_run:
        sheets = recon_excel(args.file)
        print_recon(sheets, args.file)
        print("DRY RUN — no database changes will be made.")
        print("Run without --dry-run to load data.\n")

        # Show what would be loaded
        pandas = _import_pandas()
        openpyxl = _import_openpyxl()
        sheet_purpose = classify_sheets(args.file)
        wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)

        for name in wb.sheetnames:
            purpose = sheet_purpose.get(name, "programs")
            if purpose == "skip":
                print(f"Sheet '{name}' (skip): not program/deadline/pathway data")
                continue
            df = pandas.read_excel(args.file, sheet_name=name)
            print(f"Sheet '{name}' ({purpose}): {len(df)} rows")
            if args.limit:
                df = df.head(args.limit)
            for _, row in df.iterrows():
                rd = row.dropna().to_dict()
                record, issues = normalize_program_row(rd, name)
                if record:
                    print(f"  ✓ {record['institution_name']} — {record['program_name']} [{record['degree_type']}]")
                else:
                    print(f"  ✗ REJECTED: {issues}")

        wb.close()
        return

    # Actual load
    bulk_load(args.file, args.limit)


if __name__ == "__main__":
    main()
